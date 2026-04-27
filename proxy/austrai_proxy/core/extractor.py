"""Text extraction from files — PDF, DOCX, XLSX, TXT, images.

Dependencies (PyMuPDF, python-docx, openpyxl, Pillow, pytesseract)
are included in the standard austrai install.
"""

import io
import logging
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger("austrai.extractor")


@dataclass
class ExtractionResult:
    text: str = ""
    format: str = "UNKNOWN"
    pages: int = 1
    warnings: list[str] = field(default_factory=list)


def extract_from_file(file_path: str) -> ExtractionResult:
    """Extract text from a file. Auto-detects format by extension."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Datei nicht gefunden: {file_path}")

    suffix = path.suffix.lower()
    file_bytes = path.read_bytes()

    if suffix == ".pdf":
        return _extract_pdf(file_bytes)
    elif suffix == ".docx":
        return _extract_docx(file_bytes)
    elif suffix == ".xlsx":
        return _extract_xlsx(file_bytes)
    elif suffix in (".txt", ".csv", ".md", ".json", ".xml", ".html", ".log", ".yaml", ".yml"):
        return _extract_text(file_bytes, suffix.lstrip(".").upper())
    elif suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp"):
        return _extract_image(file_bytes)
    else:
        try:
            return _extract_text(file_bytes, suffix.lstrip(".").upper())
        except Exception:
            raise ValueError(f"Dateiformat '{suffix}' nicht unterstuetzt.")


def _extract_pdf(data: bytes) -> ExtractionResult:
    try:
        import fitz
    except ImportError:
        raise ImportError("PDF-Support braucht PyMuPDF: pip install austrai")
    doc = fitz.open(stream=data, filetype="pdf")
    # PDF-Bombe abwehren: eine PDF mit 100.000 Seiten würde hier einen
    # OOM auslösen. Reale Dokumente haben selten >500 Seiten.
    MAX_PDF_PAGES = 1000
    if len(doc) > MAX_PDF_PAGES:
        doc.close()
        raise ValueError(f"PDF has too many pages ({len(doc)}). Max supported: {MAX_PDF_PAGES}")
    pages = [page.get_text() for page in doc]
    text = "\n\n".join(pages).strip()
    warnings: list[str] = []

    # Scanned / image-only PDFs have no extractable text layer. Without OCR
    # the user sees "0 entities" and wonders why — with OCR we rasterise each
    # page and run Tesseract over it. 20 characters is the threshold: real
    # PDFs almost always have more (page headers alone tend to exceed it),
    # while an "empty" scan returns e.g. only a single form-feed or footer
    # line from get_text().
    if len(text) < 20:
        try:
            import pytesseract
            from PIL import Image

            ocr_parts = []
            for page in doc:
                mat = fitz.Matrix(2, 2)  # 2x zoom improves OCR quality
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                ocr_parts.append(pytesseract.image_to_string(img, lang="deu+eng"))
            ocr_text = "\n\n".join(ocr_parts).strip()

            if ocr_text:
                text = ocr_text
                warnings.append(
                    "PDF ohne Textschicht (vermutlich gescannt) — OCR wurde "
                    "automatisch verwendet."
                )
            else:
                warnings.append(
                    "PDF enthält keinen erkennbaren Text. Falls gescannt: "
                    "Qualität prüfen oder leistungsfähigeres OCR-Tool nutzen."
                )
        except ImportError:
            warnings.append(
                "PDF ohne Textschicht erkannt (vermutlich gescannt). Tesseract-OCR "
                "ist nicht installiert — sonst könnte der Text automatisch "
                "gelesen werden. macOS: 'brew install tesseract'. Linux: "
                "'sudo apt install tesseract-ocr tesseract-ocr-deu'. Windows: "
                "https://github.com/UB-Mannheim/tesseract/wiki"
            )
        except Exception as e:
            warnings.append(f"OCR-Versuch fehlgeschlagen: {e}")

    return ExtractionResult(
        text=text,
        format="PDF",
        pages=len(doc),
        warnings=warnings,
    )


def _extract_docx(data: bytes) -> ExtractionResult:
    try:
        from docx import Document
    except ImportError:
        raise ImportError("DOCX-Support braucht python-docx: pip install austrai")
    doc = Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return ExtractionResult(text="\n".join(paragraphs), format="DOCX", pages=1)


def _extract_xlsx(data: bytes) -> ExtractionResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("XLSX-Support braucht openpyxl: pip install austrai")
    wb = load_workbook(io.BytesIO(data), read_only=True)
    parts = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"[{sheet.title}]\n" + "\n".join(rows))
    return ExtractionResult(text="\n\n".join(parts), format="XLSX", pages=len(wb.worksheets))


def _extract_text(data: bytes, fmt: str) -> ExtractionResult:
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    return ExtractionResult(text=text, format=fmt, pages=1)


def _extract_image(data: bytes) -> ExtractionResult:
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        raise ImportError("Bild-OCR braucht Tesseract + Pillow: pip install austrai")
    img = Image.open(io.BytesIO(data))
    text = pytesseract.image_to_string(img, lang="deu+eng")
    return ExtractionResult(text=text, format="IMAGE", pages=1)
